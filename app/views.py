"""
Definition of views.
"""

from datetime import datetime
from django.shortcuts import render
from django.http import HttpRequest
from Bees.Mybee import Mybee
from Bees.hive import hive
from Bees.testfunc import *
from .forms import *
import copy
from Fishes.FSS import *
from Fireflies.flyworm import go
import Ants.const as const
from Ants.draw import Draw
from Ants.model_ant import Ant
from Ants.model_field import Field
from Ants.model_const import Const
from .models import *
import openpyxl
import os
from django.http import HttpResponse

def home(request):
    """Renders the home page."""
    assert isinstance(request, HttpRequest)
    return render(request,
        'app/index.html',
        {
            'title':'Home Page',
            'year':datetime.now().year,
        })

#bees
def bees(request):
    
    funcs = {
        1: deffunc,
        2: ekli,
        3: bukin_function,
        4: cross_in_tray_function
    }
    global myhive

    if request.method == "POST":

        scoutbeecount = int(request.POST.get("scoutbeecount"))
        bestbeecount = int(request.POST.get("bestbeecount"))
        selectedbeecount = int(request.POST.get("selectedbeecount"))
        radius = int(request.POST.get("radius"))
        funcnum = int(request.POST.get("function"))
        iterationcount = int(request.POST.get("iterations"))
        
        myhive = hive(scoutbeecount,selectedbeecount,bestbeecount,3,3,[radius * 2,radius * 2],Mybee,funcs[funcnum],iterationcount)
        iterations = GiveAllIterationsForbees()
        DBSave(iterations)
        #excelsave(iterations)
        form = BeesForm()
        return render(request,
        'Bees/bees.html',
        {
            'title':'Home Page',
            'year':datetime.now().year,
            'swarm': myhive.swarm,
            'iterations': iterations,
            "form": form,
            'best': myhive.swarm[0]
        })
    else:
        assert isinstance(request, HttpRequest)
        #myhive = hive(10,5,6,3,3,[10,10],Mybee,funcs[1])
        form = BeesForm()
        return render(request,
        'Bees/bees.html',
        {
            'title':'Bees algorithm',
            'year':datetime.now().year,
            "form": form
        })

def DBSave(iterations):
    hive = BeesPopulation.objects.create(population_size = len(myhive.swarm),iteration_count = myhive.iterationcount,funcName = myhive.Function.__name__)
    
    for i in range(len(iterations)):
        for j in range(len(myhive.swarm)):
            BeeAgent.objects.create(x = iterations[i][j][0], y = iterations[i][j][1], z = iterations[i][j][2],agent_id = myhive.swarm[j].id, population_id = hive)


def GiveAllIterationsForbees():
    iterations = []
    for i in range(myhive.iterationcount):
        myhive.nextstep()
        swarm = []
        for bee in myhive.swarm:
            swarm.append(bee.getMyposition())
        iterations.append(swarm)
    return iterations


#fishes
def SaveSQL(iter, fish, weightMax, speed, function, MinX, MaxX, MinY, MaxY, error, coords):
    pop = FishesPopulation.objects.create(
        population_size = fish,
        iteration_count = iter,
        weight_max = weightMax,
        speed = speed,
        min_x = MinX,
        max_x = MaxX,
        min_y = MinY,
        max_y = MaxY,
        error = error,
        function = function)

    for i in range(fish):
        for j in range(iter):
            FishAgent.objects.create(population_id = pop,
                                      agent_id = i,
                                      x = coords[j][i].X,
                                      y = coords[j][i].Y,
                                      z = coords[j][i].Z)


def fishes(request):
    userform = FishForm()

    if request.method == "POST":
        
        userform = request.POST

        fish = int(request.POST.get("fish"))
        iter = int(request.POST.get("iterations"))
        weightMax = int(request.POST.get("weightMax"))
        speed = float(request.POST.get("speed"))
        MinX = float(request.POST.get("MinX"))
        MaxX = float(request.POST.get("MaxX"))
        MinY = float(request.POST.get("MinY"))
        MaxY = float(request.POST.get("MaxY"))
        error = float(request.POST.get("error"))
        function = int(request.POST.get("function"))

        coords = FSS.Run(iter, fish, weightMax, speed, function, MinX, MaxX, MinY, MaxY)

        SaveSQL(iter, fish, weightMax, speed, function, MinX, MaxX, MinY, MaxY, error , coords)

        MinZ = coords[iter - 1][0].Z

        for i in range(fish):
            if (coords[iter-1][i].Z < MinZ):
                MinZ = coords[iter-1][i].Z
        min = []
        for i in range(fish):
            if (coords[iter-1][i].Z <= (MinZ + error)):
                min.append(coords[iter-1][i])

        result_coords = copy.deepcopy(coords)

        kX = 690 / (MaxX - MinX)
        kY = 690 / (MaxY - MinY)

        for i in range(iter):
            for j in range(fish):
                result_coords[i][j].X = int((result_coords[i][j].X - MinX) * kX)
                result_coords[i][j].Y = int((result_coords[i][j].Y - MinY) * kY)

        if request.POST.get("DownLoad") == 'on':
            response = download(request, iter, fish, weightMax, speed, function, MinX, MaxX, MinY, MaxY, error, coords)
            return response
        else:
            result = {'positions': result_coords, 'min_result': min, "form": FishForm(userform)}
            return render(request, 'Fishes/FSS.html', context=result)

    else:
        return render(request, "Fishes/FSS.html", {"form": userform})

def download(request, iter, fish, weightMax, speed, function, MinX, MaxX, MinY, MaxY, error, coords):
    wb = openpyxl.Workbook()
    wb.create_sheet("X")
    wb.create_sheet("Y")
    wb.create_sheet("Z")
    sheet1 = wb['X']
    sheet2 = wb['Y']
    sheet3 = wb['Z']
    for i in range(iter):
        for j in range(fish):
            sheet1.cell(row=i + 1, column=j + 1).value = coords[i][j].X
            sheet2.cell(row=i + 1, column=j + 1).value = coords[i][j].Y
            sheet3.cell(row=i + 1, column=j + 1).value = coords[i][j].Z
    wb.save(str(iter) + '_' + str(fish) + '_' + str(weightMax) + '_' + str(speed) + '_' + str(function) + '_' + str(
        MinX) + '_' + str(MaxX) + '_' + str(MinY) + '_' + str(MaxY) + '_' + str(error) + '.xlsx')

    file_name = str(iter) + '_' + str(fish) + '_' + str(weightMax) + '_' + str(speed) + '_' + str(
        function) + '_' + str(
        MinX) + '_' + str(MaxX) + '_' + str(MinY) + '_' + str(MaxY) + '_' + str(error)

    if os.path.exists(file_name + '.xlsx'):
        fsock = open(file_name + '.xlsx', "rb")

        response = HttpResponse(fsock, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename =' + file_name + '.xls'

        fsock.close()
        os.remove(file_name + '.xlsx')

        return response

#fireflies
def fireflies(request):
    #form = ParamForm(request.POST or None)
    return render(request, 'Fireflies/homePage.html', locals())


def wait(request):
    if request.method == "POST": 
        nturns = int(request.POST.get('nturns', ''))
        num_worms = int(request.POST.get('num_worms', ''))
        influence_factor = float(request.POST.get('influence_factor', ''))
        max_jitter = float(request.POST.get('max_jitter', ''))
        start = float(request.POST.get('start', ''))
        end = float(request.POST.get('end', ''))
        function = request.POST.get('function', '')
        params = FireflyPopulation.objects.create(nturns = nturns, num_worms = num_worms,influence_factor=influence_factor, max_jitter=max_jitter, start=start,end=end, function=function)
    return render(request, 'Fireflies/wait.html', {'params' : [nturns, num_worms, influence_factor, max_jitter, start, end, function]})

def result(request):
    #form = ParamForm(request.POST)
    #if request.method == "POST": #and form.is_valid():
    population = FireflyPopulation.objects.last()
    results = go(population.nturns, population.num_worms, population.influence_factor, population.max_jitter, population.start, population.end, population.function, population)
    return render(request, 'Fireflies/basic.html', {'values' : ['Функция ' + population.function], 'result' : results})

def animation(request):
    return render(request, 'Fireflies/animation.html')

#ants
def init(request):
    const = Const(id = 0)
    const.save()
    return HttpResponse('OK')

def ants(request):
    #nconst = Const.objects.get(id = 0)
    #userForm = AntsForm(request.GET, instance = nconst)
    userForm = AntsForm()
    if userForm.is_valid():
        userForm.save()
    global antsfield
    antsfield = Field()
    antsfield.initField()
    antsfield.createAnts()
    return render(request, 'Ants/aco.html', {'form':userForm})

def getNextMove(request):
    draw = Draw()
    antsfield.moveAnts()
    matrix = draw.getDrawingMatrix(antsfield)
    antsfield.globalEvaporate()
    return render(request,
        'Ants/data.html',
        {
            'matrix': matrix
        })
